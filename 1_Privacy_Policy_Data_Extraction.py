from selenium import webdriver
import time
import xlrd
import json
from bs4 import BeautifulSoup
import openpyxl
import urllib.request
from bs4.element import Comment
import pandas as pd
from io import open
import os
import re

driver = webdriver.Chrome('./chromedriver')
driver.maximize_window()

website_url = ''
url_content = ''

df = pd.read_excel('./urls.xlsx', sheet_name='url')
df = df.fillna('')


col = 1
for index, row in df.iterrows():
    col += 1
    if row['Privacy Policy URL'] != '':
        continue
    if row['weburl'] == '':
        continue
    website_url = 'http://' + row['weburl']
    print(website_url)
    driver.set_page_load_timeout(60)
    try:
    	driver.get(website_url)
    	print('1')
    except:
        print('Skipped')
        workbook = openpyxl.load_workbook('./urls.xlsx')
        sheet = workbook["url"]
        count = "B" + str(col)
        sheet[count].value = 'Skipped'
        workbook.save('./urls.xlsx')
        continue

    time.sleep(3)
    print('2')
    try:
        driver.execute_script("arguments[0].click();",
                              driver.find_element_by_xpath("""//*[contains(text(),'Privacy Policy')]"""))
        print('found1')
    except:
        try:
            driver.execute_script("arguments[0].click();",
                                  driver.find_element_by_xpath("""//*[contains(text(),'Privacy')]"""))
            print('found2')
        except:
            try:
                driver.execute_script("arguments[0].click();",
                                      driver.find_element_by_xpath("""//*[contains(text(),'privacy')]"""))
                print('found3')
            except:
                print('No Privacy Policy')
                workbook = openpyxl.load_workbook('./urls.xlsx')
                sheet = workbook["url"]
                count = "B" + str(col)
                sheet[count].value = 'No Privacy Policy'
                workbook.save('./urls.xlsx')
                continue
    time.sleep(4)
    print('3')
    url = driver.current_url
    workbook = openpyxl.load_workbook('./urls.xlsx')
    sheet = workbook["url"]
    count = "B" + str(col)
    sheet[count].value = url
    workbook.save('./urls.xlsx')
    r = driver.page_source
    with open(
            ".//PolicyData//HTMLFiles//" +
            row["weburl"]+ ".html", 'w', encoding="utf-8") as f:
        f.write(r)
